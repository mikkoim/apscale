import subprocess, gzip, datetime, os, subprocess, pickle, glob, openpyxl, shutil, psutil
import pandas as pd
import numpy as np
from Bio import SeqIO
from joblib import Parallel, delayed
from pathlib import Path
from Bio.SeqIO.FastaIO import SimpleFastaParser
from io import StringIO
from tqdm import tqdm
from openpyxl.utils.dataframe import dataframe_to_rows
from functools import reduce
import re

## clustering function to cluster all sequences in input fasta with given pct_id
def otu_clustering(project = None, comp_lvl = None, cores = None, pct_id = None):
    """Function to apply OTU clustering to a given gzipped file. Outputs a fasta file
    with all centroid sequences."""

    ## define the name for the output fasta
    ## create an output path to write to
    sample_name_out_1 = 'OTUs_with_chimeras.fasta.gz'
    output_path = Path(project).joinpath('7_otu_clustering', 'data', sample_name_out_1)

    ## give user output
    print('{}: Starting OTU clustering. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))

    ## reduce cores to 75% of available ressources to prevent overheating while clustering / denoising:
    if cores > int(psutil.cpu_count() * 0.75):
        cores = int(psutil.cpu_count() * 0.75)

    ## run vsearch --cluster_size to cluster OTUs
    ## use --log because for some reason no info is written to stderr with this command
    ## write stdout to uncompressed output at runtime
    with open(output_path.with_suffix(''), 'w') as output:
        f = subprocess.run(['vsearch',
                            '--cluster_size', str(Path(project).joinpath('6_dereplication_pooling', 'data', 'pooling', 'pooled_sequences_dereplicated.fasta.gz')),
                            '--id', str(pct_id / 100),
                            '--sizein', '--sizeout', '--relabel', 'OTU_',
                            '--centroids', '-', '--fasta_width', str(0), '--quiet',
                            '--log', str(Path(project).joinpath('7_otu_clustering', 'temp', 'clustering_log.txt')),
                            '--threads', str(cores)], stdout = output)

    ## compress the output, remove uncompressed output
    with open(output_path.with_suffix(''), 'rb') as in_stream, gzip.open(output_path, 'wb', comp_lvl) as out_stream:
            shutil.copyfileobj(in_stream, out_stream)
    os.remove(output_path.with_suffix(''))

    ## collect processed and passed reads from the log file
    with open(Path(project).joinpath('7_otu_clustering', 'temp', 'clustering_log.txt')) as log_file:
        content = log_file.read()
        seqs = re.search(r'(\d+)(?= sequences)', content).group(0)
        clusters = re.search(r'(?<=Clusters: )(\d+)', content).group(0)
        version = re.search(r'[^,]*', content).group(0)
        
        finished = '{}'.format(datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    print('{}: Clustered unique {} sequences into {} OTUs.'.format(datetime.datetime.now().strftime("%H:%M:%S"), seqs, clusters))
    print('{}: Starting chimera removal from the OTUs. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))

    ## run vsearch --uchime_denovo to remove chimeric sequences from the OTUs
    f = subprocess.run(['vsearch',
                        '--uchime_denovo', str(Path(project).joinpath('7_otu_clustering', 'data', sample_name_out_1)),
                        '--relabel', 'OTU_',
                        '--nonchimeras', str(Path(project).joinpath('7_otu_clustering', '{}_OTUs.fasta'.format(Path(project).stem))),
                        '-fasta_width', str(0), '--quiet'])

    ## collect processed and passed reads from the output fasta, since it is not reported in the log
    f = list(SimpleFastaParser(open(Path(project).joinpath('7_otu_clustering', '{}_OTUs.fasta'.format(Path(project).stem)))))
    print('{}: {} chimeras removed from {} OTU sequences.'.format(datetime.datetime.now().strftime("%H:%M:%S"), int(clusters) - len(f), clusters))
    print('{}: OTUs saved to {}.'.format(datetime.datetime.now().strftime("%H:%M:%S"), Path(project).joinpath('7_otu_clustering', '{}_OTUs.fasta'.format(Path(project).stem))))

## remapping function to remap the individual reads to the OTUs via vsearch
def remapping(file, project = None, pct_id = None):
    """Function to remap the sequences of a dereplicated file against the OTU list
    as database."""

    ## extract the sample name from the file name for the otu table
    sample_name_out = '{}'.format(Path(file).with_suffix('').with_suffix('').name).replace('_PE_trimmed_filtered_dereplicated', '')

    ## run vsearch --usearch_global to remap the individual files vs the generated
    ## OTU fasta, capture log and directly pickle the output as dataframe for read table generation
    f = subprocess.run(['vsearch',
                        '--usearch_global', str(Path(file)),
                        '--db', str(Path(project).joinpath('7_otu_clustering', '{}_OTUs.fasta'.format(Path(project).stem))),
                        '--id', str(pct_id / 100),
                        '--output_no_hits',
                        '--maxhits', '1',
                        '--otutabout', '-', '--quiet', '--threads', str(1),
                        '--log', str(Path(project).joinpath('7_otu_clustering', 'temp', '{}_mapping_log.txt'.format(sample_name_out)))], capture_output = True)

    ## directly parse the output to a pandas dataframe
    otu_tab = pd.read_csv(StringIO(f.stdout.decode('ascii', errors = 'ignore')), sep = '\t')

    ## handle empty outputs correctly
    if not otu_tab.empty:
        otu_tab = otu_tab.set_axis(['ID', sample_name_out], axis = 1, inplace = False)
    else:
        otu_tab[sample_name_out] = ""

    ## collect processed and mapped reads from output and pickle them to data logs
    with open(Path(project).joinpath('7_otu_clustering', 'temp', '{}_mapping_log.txt'.format(sample_name_out))) as log_file:
        content = log_file.read().split('\n')
        version = content[0].split(',')[0]
        seqs, mapped = content[5].split(' ')[6], content[5].split(' ')[4]
        finished = '{}'.format(datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    ## give user output
    print('{}: {}: {} of {} sequences mapped successfully.'.format(datetime.datetime.now().strftime("%H:%M:%S"), sample_name_out, mapped, seqs))

    ## pickle log data first for log generation
    with open(Path(project).joinpath('7_otu_clustering', 'temp', '{}_log.pkl'.format(sample_name_out)), 'wb') as log:
        pickle.dump([sample_name_out, finished, version, seqs, mapped], log)

    ## pickle otu tab dataframes for otu table generation
    with open(Path(project).joinpath('7_otu_clustering', 'temp', '{}_otu_tab.pkl'.format(sample_name_out)), 'wb') as log:
        pickle.dump(otu_tab, log)

## main function for the otu clustering
def main(project = Path.cwd()):
    """Main function of the script. Default values can be changed via the input file.
    Will cluster OTUs, perform chimera removal, remap the individual files and
    generate an OTU table."""

    ## create temporal output folder
    try:
        os.mkdir(Path(project).joinpath('7_otu_clustering', 'temp'))
    except FileExistsError:
        pass

    ## collect variables from the settings file
    gen_settings = pd.read_excel(Path(project).joinpath('Settings.xlsx'), sheet_name = '0_general_settings')
    cores, comp_lvl = gen_settings['cores to use'].item(), gen_settings['compression level'].item()

    settings = pd.read_excel(Path(project).joinpath('Settings.xlsx'), sheet_name = '7_otu_clustering')
    pct_id, to_excel = settings['pct id'].item(), settings['to excel'].item()

    ## run OTU clustering function
    otu_clustering(project = project, comp_lvl = comp_lvl, cores = cores, pct_id = pct_id)

    ## gather files for remapping of OTUS
    input = glob.glob(str(Path(project).joinpath('6_dereplication_pooling', 'data', 'dereplication', '*.fasta.gz')))

    print('{}: Starting to remap {} input files.'.format(datetime.datetime.now().strftime("%H:%M:%S"), len(input)))

    ## run remapping parallelized to speed up the process
    Parallel(n_jobs = cores)(delayed(remapping)(file, project = project, pct_id = pct_id) for file in input)

    ## write log for the clustering from pkl logs
    summary_logs = glob.glob(str(Path(project).joinpath('7_otu_clustering', 'temp', '*_log.pkl')))
    summary = [pickle.load(open(line, 'rb')) for line in summary_logs]

    log_df = pd.DataFrame(summary, columns = ['File', 'finished at', 'program version', 'processed sequences', 'mapped sequences'])
    log_df = log_df.sort_values(by = 'File')
    log_df.to_excel(Path(project).joinpath('7_otu_clustering', 'Logfile_7_otu_clustering.xlsx'),
                    index = False,
                    sheet_name = '7_otu_clustering')

    ## add log to the project report
    wb = openpyxl.load_workbook(Path(project).joinpath('Project_report.xlsx'))
    writer = pd.ExcelWriter(Path(project).joinpath('Project_report.xlsx'), engine = 'openpyxl')
    writer.book = wb

    ## write the output to excel
    log_df.to_excel(writer, sheet_name = '7_otu_clustering', index = False)
    wb.save(Path(project).joinpath('Project_report.xlsx'))
    writer.close()

    ## generate OTU table, first extract all OTUs and sequences from fasta file
    otu_list = list(SimpleFastaParser(open(Path(project).joinpath('7_otu_clustering', '{}_OTUs.fasta'.format(Path(project).stem)))))
    otu_table = pd.DataFrame(otu_list, columns = ['ID', 'Seq'])
    seq_col = otu_table.pop('Seq')

    ## extract individual OTU tabs from the clustering output, rename columns correctly, merge all individual tabs with the otu table frame
    otu_tabs = glob.glob(str(Path(project).joinpath('7_otu_clustering', 'temp', '*_otu_tab.pkl')))
    otu_tabs = [pickle.load(open(tab_file, 'rb')) for tab_file in otu_tabs]
    otu_tabs = [tab.rename(columns = {tab.columns[0] : 'ID'}) for tab in otu_tabs]
    otu_tabs = [pd.merge(otu_table, tab, on = 'ID', how = 'outer').set_index('ID') for tab in tqdm(otu_tabs, desc = 'Generating OTU table')]

    ## collapse all individual dataframes into the OTU table, replace nan values with 0, reset index
    otu_table = pd.concat(otu_tabs, axis = 1)
    otu_table = otu_table.reset_index(level = 0).fillna(0)
    otu_table = pd.concat([otu_table[['ID']], otu_table[otu_table.columns.difference(['ID'])].sort_index(axis = 1)], ignore_index = False, axis = 1)

    ## move sequences to the end of the dataframe
    otu_table.insert(len(otu_table.columns), 'Seq', seq_col)

    ## save the final OTU table if option is selected
    if to_excel:
        wb = openpyxl.Workbook(write_only = True)
        ws = wb.create_sheet('OTU table')

        ## save the output line by line for optimized memory usage
        for row in tqdm(dataframe_to_rows(otu_table, index = False, header = True),
                                          total = len(otu_table.index),
                                          desc = '{}: Lines written to OTU table'.format(datetime.datetime.now().strftime("%H:%M:%S")),
                                          unit = ' lines'):
            ws.append(row)

        ## save the output (otu table)
        print('{}: Saving the OTU table to excel. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))
        wb.save(Path(project).joinpath('7_otu_clustering', '{}_OTU_table.xlsx'.format(Path(project).stem)))
        wb.close()
        print('{}: OTU table saved to {}.'.format(datetime.datetime.now().strftime("%H:%M:%S"), Path(project).joinpath('7_otu_clustering', '{}_OTU_table.xlsx'.format(Path(project).stem))))

    ## save to parquet if selected, compress with snappy
    print('{}: Saving the OTU table to parquet. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))
    otu_table.to_parquet(Path(project).joinpath('7_otu_clustering', '{}_OTU_table.parquet.snappy'.format(Path(project).stem)), index = False)
    print('{}: OTU table saved to {}.'.format(datetime.datetime.now().strftime("%H:%M:%S"), Path(project).joinpath('7_otu_clustering', '{}_OTU_table.parquet.snappy'.format(Path(project).stem))))

    ## remove temporary files
    shutil.rmtree(Path(project).joinpath('7_otu_clustering', 'temp'))

if __name__ == "__main__":
    main()
